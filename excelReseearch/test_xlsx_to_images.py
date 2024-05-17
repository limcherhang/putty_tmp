from openpyxl import load_workbook

# 打开xlsx文件
wb = load_workbook('test.xlsx')

# 遍历每个工作表
for sheet_name in wb.sheetnames:
    # 选择当前工作表
    ws = wb[sheet_name]

    # 遍历当前工作表中的每个图片
    for img in ws._images:
        # 获取图片的放置位置
        anchor = img.anchor
        from_col = anchor._from.col
        from_row = anchor._from.row
        to_col = anchor.to.col
        to_row = anchor.to.row
        print(f"Image placed from Cell: {from_col}, {from_row} to Cell: {to_col}, {to_row}")

        # 获取图像的文件名
        image_filename = f'./output_images/{sheet_name}_{img._id}.{img.format}'

        # 保存图像到本地文件
        with open(image_filename, 'wb') as f:
            f.write(img._data())

# 关闭xlsx文件
wb.close()